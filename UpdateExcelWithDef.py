# importing openpyxl module
import openpyxl as xl


def feed_input_file(newDataFile):

    # open the data file from which new data to be updated
    #filename ="NewData.xlsx"

    # Here I am trying to pass parameter value, should be inputted  as our data file to update our file having template for powerpoint presentation file
    Workbook1 = xl.load_workbook(newDataFile)
    Worksheet1 = Workbook1.worksheets[0]

# Open the file for powerpoint updation
    Updatefile ="Update Chart.xlsx"
    Workbook2 = xl.load_workbook(Updatefile)
    Worksheet2 = Workbook2.worksheets[0]

# Count total number of rows in new datafile
    CountRow = Worksheet1.max_row

# Count total number of columns in new datafile
    CountColumn = Worksheet1.max_column

#Reading and writing of new data file to data file linked with our Presentation
    for i in range (1, CountRow + 1):
	    for j in range (1, CountColumn + 1):
		# reading cell value from new data file
		    ReadCell = Worksheet1.cell(row = i, column = j)

		# writing the read value to excel file linked with our Presentation file
		    Worksheet2.cell(row = i, column = j).value = ReadCell.value

#Save the file link with our presentation file
    Workbook2.save(Updatefile)

# x="NewData.xlsx"
#
# feed_input_file(x)

#import pandas as pd
#import os

#This is also causing same issue of hard coding of the input file
# for filename in os.listdir(os.getcwd()):
#    if filename.endswith("*wData.xlsx"):
#     print(filename)
#     #do your operation
#     data_df = pd.read_excel(filename, sheet_name=None)
#     os.remove(filename)
